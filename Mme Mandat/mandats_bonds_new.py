import pandas as pd
import numpy as np
import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from faker import Faker

fake = Faker()

# Données pour la génération aléatoire
notes_credit = [
    # Investment Grade
    "AAA", "AA+", "AA", "AA-", "A+", "A", "A-", "BBB+", "BBB", "BBB-",
    # High Yield
    "BB+", "BB", "BB-", "B+", "B", "B-", "CCC+", "CCC", "CCC-"
]

notes_moodys = ["Aaa", "Aa1", "Aa2", "Aa3", "A1", "A2", "A3", "Baa1", "Baa2", "Baa3", 
                "Ba1", "Ba2", "Ba3", "B1", "B2", "B3", "Caa1", "Caa2", "Caa3"]

notes_fitch = ["AAA", "AA+", "AA", "AA-", "A+", "A", "A-", "BBB+", "BBB", "BBB-",
               "BB+", "BB", "BB-", "B+", "B", "B-", "CCC+", "CCC", "CCC-"]

notes_lmdg = ["A+", "A", "A-", "BBB+", "BBB", "BBB-", "BB+", "BB", "BB-", "B+", "B", "B-"]

sectors = ['Finance', 'Technology', 'Healthcare', 'Consumer', 'Industrial', 'Energy', 'Materials', 'Utilities']
countries = ['FR', 'DE', 'IT', 'ES', 'NL', 'BE', 'US', 'GB']
payment_ranks = ['Senior', 'Subordinated', 'Junior Subordinated']
currencies = ['EUR', 'USD', 'GBP']
strategy_codes = ['CONV', 'DIR', 'ARB', 'STRAT']
genres_titre = ['Obligation', 'Obligation Convertible', 'Obligation Perpétuelle']
conditions_portefeuille = [
    "Pas d'investissement dans le pétrole",
    "Pas d'armement",
    "Minimum 50% Investment Grade",
    "Maximum 30% High Yield",
    "Pas de tabac",
    "ESG Score minimum B",
    "Maximum 20% USD"
]

# Liste des prénoms pour les mandats
noms_mandats = [
    "Sophie", "Emma", "Charlotte", "Alice", "Marie",
    "Pierre", "Louis", "Thomas", "Nicolas", "Antoine"
]

# Créer un nouveau workbook
wb = Workbook()
ws = wb.active

# Définir les en-têtes
headers = [
    "PORTEFEUILLE",
    "STRATEGY CODE",
    "ISIN",
    "GENRE DU TITRE",
    "DATE SYSTEME",
    "QUANTITE",
    "DATE DERNIER COURS",
    "VALORISATION DEV PTF",
    "% VALORISATION TOT",
    "% + - VALUE",
    "Date",
    "Condition",
    "Stratégie",
    "name",
    "security name",
    "cpn",
    "INDUSTRY_SECTOR",
    "currency",
    "maturity",
    "is perpetual",
    "payment rank",
    "TICKER",
    "RTG_SP_NO_WATCH",
    "RTG_MOODY_NO_WATCH",
    "country",
    "Contrinution To Yield",
    "Contribution to duration",
    "YAS_MOD_DUR",
    "YAS_ASW_SPREAD",
    "NXT_CALL_DT",
    "Final Maturity",
    "YAS_BOND_YLD",
    "Yield",
    "Moody's",
    "S&P",
    "Fitch",
    "LMDG",
    "rating class",
    "worst rating",
    "Worst Raitng F",
    "Ratingtimesweights",
    "Px_Last",
    "Valeur Marchande",
    "Poids",
    "CouponContribution",
    "CoCo",
    "Avg Rating Portefeuille",
    "Ca",
    "Condition",
    "Référencement",
    "Univers Investissable",
    "Nom du Mandats"
]

# Style pour les en-têtes
header_font = Font(bold=True)
header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

# Ajouter les en-têtes
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill

# Générer des données aléatoires
row = 2
mandats = list(range(1, len(noms_mandats) + 1))  # Liste des mandats

def generate_random_bonds(num_bonds):
    bonds = []
    
    for _ in range(num_bonds):
        # Dates
        date_systeme = datetime.now()
        date_dernier_cours = date_systeme - timedelta(days=random.randint(0, 5))
        maturity_date = fake.date_between(start_date='+1y', end_date='+10y')
        has_call = random.choice([True, False])
        call_date = fake.date_between(start_date='+6m', end_date=maturity_date) if has_call else None
        
        # Ratings
        sp_rating = random.choice(notes_credit)
        moodys_rating = random.choice(notes_moodys)
        fitch_rating = random.choice(notes_fitch)
        lmdg_rating = random.choice(notes_lmdg)
        
        # Financials
        coupon = round(random.uniform(0.5, 8.0), 2)
        quantite = round(random.uniform(100000, 10000000), 0)
        px_last = round(random.uniform(85.0, 115.0), 2)
        valeur_marchande = quantite * px_last / 100
        poids = round(random.uniform(0.1, 5.0), 2)
        
        # Sector et payment rank
        sector = random.choice(sectors)
        payment_rank_val = random.choice(payment_ranks)
        is_coco = "Yes" if sector == "Finance" and payment_rank_val == "Junior Subordinated" else "No"
        
        # Yields et Duration
        yas_bond_yld = round(random.uniform(1.0, 10.0), 2)
        yas_mod_dur = round(random.uniform(0.5, 12.0), 2)
        yas_asw_spread = round(random.uniform(50, 500), 0)
        contribution_to_yield = round(yas_bond_yld * (poids/100), 4)
        contribution_to_duration = round(yas_mod_dur * (poids/100), 4)
        
        bond = {
            'PORTEFEUILLE': "",  # Sera rempli plus tard
            'STRATEGY CODE': random.choice(strategy_codes),
            'ISIN': fake.bothify(text='??#########'),
            'GENRE DU TITRE': random.choice(genres_titre),
            'DATE SYSTEME': date_systeme.strftime('%Y-%m-%d'),
            'QUANTITE': quantite,
            'DATE DERNIER COURS': date_dernier_cours.strftime('%Y-%m-%d'),
            'VALORISATION DEV PTF': valeur_marchande,
            '% VALORISATION TOT': poids,
            '% + - VALUE': round(random.uniform(-15, 15), 2),
            'Date': date_systeme.strftime('%Y-%m-%d'),
            'Condition': random.choice(conditions_portefeuille),
            'Stratégie': random.choice(strategy_codes),
            'name': fake.company() + ' ' + str(maturity_date.year),
            'security name': fake.company() + ' Bond ' + str(maturity_date.year),
            'cpn': coupon,
            'INDUSTRY_SECTOR': sector,
            'currency': random.choice(currencies),
            'maturity': maturity_date.strftime('%Y-%m-%d'),
            'is perpetual': "Yes" if "Perpétuelle" in genres_titre else "No",
            'payment rank': payment_rank_val,
            'TICKER': fake.bothify(text='????'),
            'RTG_SP_NO_WATCH': sp_rating,
            'RTG_MOODY_NO_WATCH': moodys_rating,
            'country': random.choice(countries),
            'Contrinution To Yield': contribution_to_yield,
            'Contribution to duration': contribution_to_duration,
            'YAS_MOD_DUR': yas_mod_dur,
            'YAS_ASW_SPREAD': yas_asw_spread,
            'NXT_CALL_DT': call_date.strftime('%Y-%m-%d') if call_date else "",
            'Final Maturity': maturity_date.strftime('%Y-%m-%d'),
            'YAS_BOND_YLD': yas_bond_yld,
            'Yield': yas_bond_yld,
            'Moody\'s': moodys_rating,
            'S&P': sp_rating,
            'Fitch': fitch_rating,
            'LMDG': lmdg_rating,
            'rating class': "IG" if sp_rating.startswith(('AAA', 'AA', 'A', 'BBB')) else "HY",
            'worst rating': min(sp_rating, fitch_rating, key=lambda x: notes_credit.index(x)),
            'Worst Raitng F': min(sp_rating, fitch_rating, key=lambda x: notes_credit.index(x)),
            'Ratingtimesweights': round(random.uniform(0.1, 5.0), 2),
            'Px_Last': px_last,
            'Valeur Marchande': valeur_marchande,
            'Poids': poids,
            'CouponContribution': round(coupon * (poids/100), 4),
            'CoCo': is_coco,
            'Avg Rating Portefeuille': random.choice(notes_credit),
            'Ca': round(random.uniform(-2, 2), 2),
            'Condition': random.choice(conditions_portefeuille),
            'Référencement': random.choice(["Yes", "No"]),
            'Univers Investissable': random.choice(["Yes", "No"]),
            'Nom du Mandats': ""  # Sera rempli plus tard
        }
        bonds.append(bond)
    
    return bonds

for mandat in mandats:
    # Nombre aléatoire de lignes entre 20 et 50 pour ce mandat
    nb_lignes = random.randint(20, 50)
    
    # Répartition entre obligations, monétaire et cash
    nb_monetaire = random.randint(1, 3)  # 1 à 3 lignes de monétaire
    nb_cash = random.randint(1, 2)       # 1 à 2 lignes de cash
    nb_obligations = nb_lignes - nb_monetaire - nb_cash
    
    # Générer les lignes pour ce mandat
    types_lignes = (
        ['obligation'] * nb_obligations +
        ['monetaire'] * nb_monetaire +
        ['cash'] * nb_cash
    )
    random.shuffle(types_lignes)  # Mélanger l'ordre des lignes
    
    for type_ligne in types_lignes:
        if type_ligne == 'obligation':
            bond = generate_random_bonds(1)[0]
            row_data = [
                mandat, bond['STRATEGY CODE'], bond['ISIN'], bond['GENRE DU TITRE'], bond['DATE SYSTEME'], bond['QUANTITE'],
                bond['DATE DERNIER COURS'], bond['VALORISATION DEV PTF'], bond['% VALORISATION TOT'], bond['% + - VALUE'],
                bond['Date'], bond['Condition'], bond['Stratégie'], bond['name'], bond['security name'], bond['cpn'],
                bond['INDUSTRY_SECTOR'], bond['currency'], bond['maturity'], bond['is perpetual'], bond['payment rank'],
                bond['TICKER'], bond['RTG_SP_NO_WATCH'], bond['RTG_MOODY_NO_WATCH'], bond['country'], bond['Contrinution To Yield'],
                bond['Contribution to duration'], bond['YAS_MOD_DUR'], bond['YAS_ASW_SPREAD'], bond['NXT_CALL_DT'], bond['Final Maturity'],
                bond['YAS_BOND_YLD'], bond['Yield'], bond['Moody\'s'], bond['S&P'], bond['Fitch'], bond['LMDG'], bond['rating class'],
                bond['worst rating'], bond['Worst Raitng F'], bond['Ratingtimesweights'], bond['Px_Last'], bond['Valeur Marchande'],
                bond['Poids'], bond['CouponContribution'], bond['CoCo'], bond['Avg Rating Portefeuille'], bond['Ca'], bond['Condition'],
                bond['Référencement'], bond['Univers Investissable'], noms_mandats[mandat-1]
            ]
        elif type_ligne == 'monetaire':
            # Sélectionner un fonds monétaire
            isin, nom = random.choice([("FR0010510800", "Amundi Cash Institution SRI"), ("FR0000970251", "AXA Court Terme")])
            date_systeme = datetime.now()
            date_dernier_cours = date_systeme - timedelta(days=random.randint(0, 2))
            quantite = random.randint(1000000, 5000000)
            px_last = round(random.uniform(98.0, 102.0), 2)
            valeur_marchande = quantite * px_last / 100
            poids = round(random.uniform(5, 15), 2)
            
            row_data = [
                mandat,                     # PORTEFEUILLE
                "MON",                      # STRATEGY CODE
                isin,                       # ISIN
                "Monétaire",               # GENRE DU TITRE
                date_systeme.strftime('%Y-%m-%d'),  # DATE SYSTEME
                quantite,                   # QUANTITE
                date_dernier_cours.strftime('%Y-%m-%d'),  # DATE DERNIER COURS
                valeur_marchande,           # VALORISATION DEV PTF
                poids,                      # % VALORISATION TOT
                round(random.uniform(-1, 1), 2),  # % + - VALUE
                date_systeme.strftime('%Y-%m-%d'),  # Date
                random.choice(conditions_portefeuille),  # Condition
                "MON",                      # Stratégie
                nom,                        # name
                nom + " Fund",              # security name
                round(random.uniform(2, 3), 2),  # cpn
                "Finance",                  # INDUSTRY_SECTOR
                "EUR",                      # currency
                "",                         # maturity
                "No",                       # is perpetual
                "Senior",                   # payment rank
                isin[:4],                   # TICKER
                "A-1+",                     # RTG_SP_NO_WATCH
                "P-1",                      # RTG_MOODY_NO_WATCH
                "FR",                       # country
                round(random.uniform(0.01, 0.05), 4),  # Contrinution To Yield
                round(random.uniform(0.01, 0.05), 4),  # Contribution to duration
                round(random.uniform(0.1, 0.5), 2),    # YAS_MOD_DUR
                "",                         # YAS_ASW_SPREAD
                "",                         # NXT_CALL_DT
                "",                         # Final Maturity
                round(random.uniform(2, 3), 2),  # YAS_BOND_YLD
                round(random.uniform(2, 3), 2),  # Yield
                "P-1",                      # Moody's
                "A-1+",                     # S&P
                "F1+",                      # Fitch
                "A+",                       # LMDG
                "IG",                       # rating class
                "A-1+",                     # worst rating
                "A-1+",                     # Worst Raitng F
                poids,                      # Ratingtimesweights
                px_last,                    # Px_Last
                valeur_marchande,           # Valeur Marchande
                poids,                      # Poids
                round(random.uniform(0.01, 0.05), 4),  # CouponContribution
                "No",                       # CoCo
                "A-1+",                     # Avg Rating Portefeuille
                0,                          # Ca
                random.choice(conditions_portefeuille),  # Condition
                "Yes",                      # Référencement
                "Yes",                      # Univers Investissable
                noms_mandats[mandat-1]      # Nom du Mandats
            ]
        else:  # cash
            date_systeme = datetime.now()
            date_dernier_cours = date_systeme
            quantite = random.randint(500000, 2000000)
            poids = round(random.uniform(2, 8), 2)
            
            row_data = [
                mandat,                     # PORTEFEUILLE
                "CASH",                     # STRATEGY CODE
                "",                         # ISIN
                "Cash",                     # GENRE DU TITRE
                date_systeme.strftime('%Y-%m-%d'),  # DATE SYSTEME
                quantite,                   # QUANTITE
                date_dernier_cours.strftime('%Y-%m-%d'),  # DATE DERNIER COURS
                quantite,                   # VALORISATION DEV PTF
                poids,                      # % VALORISATION TOT
                0,                          # % + - VALUE
                date_systeme.strftime('%Y-%m-%d'),  # Date
                random.choice(conditions_portefeuille),  # Condition
                "CASH",                     # Stratégie
                "Cash",                     # name
                "Cash Position",            # security name
                0,                          # cpn
                "",                         # INDUSTRY_SECTOR
                "EUR",                      # currency
                "",                         # maturity
                "No",                       # is perpetual
                "",                         # payment rank
                "",                         # TICKER
                "",                         # RTG_SP_NO_WATCH
                "",                         # RTG_MOODY_NO_WATCH
                "",                         # country
                0,                          # Contrinution To Yield
                0,                          # Contribution to duration
                0,                          # YAS_MOD_DUR
                "",                         # YAS_ASW_SPREAD
                "",                         # NXT_CALL_DT
                "",                         # Final Maturity
                0,                          # YAS_BOND_YLD
                0,                          # Yield
                "",                         # Moody's
                "",                         # S&P
                "",                         # Fitch
                "",                         # LMDG
                "",                         # rating class
                "",                         # worst rating
                "",                         # Worst Raitng F
                0,                          # Ratingtimesweights
                100,                        # Px_Last
                quantite,                   # Valeur Marchande
                poids,                      # Poids
                0,                          # CouponContribution
                "No",                       # CoCo
                "",                         # Avg Rating Portefeuille
                0,                          # Ca
                random.choice(conditions_portefeuille),  # Condition
                "Yes",                      # Référencement
                "Yes",                      # Univers Investissable
                noms_mandats[mandat-1]      # Nom du Mandats
            ]
        
        # Ajouter la ligne de données
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=value)
        
        row += 1

# Ajuster la largeur des colonnes
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Sauvegarder le fichier Excel
wb.save('mandats.xlsx')
